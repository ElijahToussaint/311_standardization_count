# imports
import os
import pandas
import openpyxl
import datetime
from openpyxl import load_workbook


# function that iterates through the columns (first row) of an excel file
# the function returns an int (column cell value position)
def selectColumn(columnName):
    result = None
    # create a list to store column values
    placeList = []
    # iterate through the columns (first row) and store there values
    for column in ws.iter_cols(max_row=1, values_only=True):
        for cell in column:
            placeList.append(cell)
    places = placeList
    # iterate through the column values and save the requested column cell
    for place in places:
        if place == columnName:
            result = placeList.index(place)
            break
    if result == None:
        print('%s does not exist...' % columnName)
    else:
        print(placeList[result])
    return result


# function that gets the cell values of an entire column in an excel file
# the function returns a list of strings (cell values)
def getColumn(columnName):
    if selectColumn(columnName) != None:
        # variable that will store the column value
        # +1 is added to the variable because min_col and max_col has a range of cell value +1
        columnValue = selectColumn(columnName) + 1
        # list that will store the cell values of the selected column
        cellList = []
        # iterate through the selected column and stores its cell values
        for column in ws.iter_cols(min_col=columnValue, max_col=columnValue, min_row=2, values_only=True):
            for cell in column:
                if cell != None:
                    cellList.append(cell.strip())
                else:
                    cellList.append(cell)
        # prints to console the cells that are stored in the list
        cells = cellList
        '''
        for cell in cells:
            if cell != None:
                print(cell)
        '''
        result = cells
    else:
        result = None
        print('%s does not exist...' % columnName)
    return result


# function the matches two columns (cell values) and adds them to a dictionary
# the function returns a list of dictionaries (dictionaries contains strings)
# the dictionary values are stored in this format -> {column1 : column2} -> column1 = column2
def matchColumn(column1, column2):
    # variables that store the list of columns by calling the getColumn() function
    firstColumn = getColumn(column1)
    secondColumn = getColumn(column2)
    if firstColumn == None or secondColumn == None:
        result = None
        print('Columns could not be matched.')
    else:
        # list that stores the column values (strings) as row values (ints)
        firstColumnValues = []
        secondColumnValues = []
        # list that stores the dictionary values of both columns
        dictsList = []
        # iterates through the first column and store the values as rows (ints)
        for row in firstColumn:
            if row != None:
                firstColumnValue = firstColumn.index(row)
                firstColumnValues.append(firstColumnValue)
        # iterates through the second column and store the values as rows (ints)
        for row in secondColumn:
            if row != None:
                secondColumnValue = secondColumn.index(row)
                secondColumnValues.append(secondColumnValue)
        # stores the row values (ints) as a list
        # converts the lists to sets and merges them (removing duplicate values)
        matchList = list(
            set(firstColumnValues).intersection(secondColumnValues))
        matchList.sort()
        # print(matchList)
        # iterates through the list of row values
        # converts the rows values (ints) back into strings
        # stores the the values as a list of dictionaries (strings)
        for value in matchList:
            print(firstColumn[value] + ' = ' + secondColumn[value])
            dicts = {firstColumn[value]: secondColumn[value]}
            dictsList.append(dicts)
        result = dictsList
    # print(result)
    return result


# function that locates a specific directory
# returns the relative path of the directory
def findDirectory(city):
    result = None
    path = './raw_data/311_raw/'
    for directories in os.listdir(path):
        relativeDirectories = os.path.join(path, directories)
        if os.path.isdir(relativeDirectories):
            # print(directories)
            if directories == city:
                # print(city)
                result = os.path.join(path, city)
                break
            else:
                pass
        else:
            print('There are no directories in this folder.')
            break
    if result == None:
        print(city + ' does not exist in directory.')
    else:
        print(result)
    return result


# fuction that list the files(csv) in a specific directory
# returns an array of files
def listFiles(directoryPath):
    result = []
    relativeFiles = os.listdir(directoryPath)
    for files in relativeFiles:
        file = files.split('.')
        filename = file[0]
        extension = file[1]
        if extension == 'csv':
            # print(files)
            result.append(files)
    result.sort()
    if not result:
        print('There are no files in directory')
    else:
        for files in result:
            print(files)
    return result


# function that locates a specific file
# returns the relative path of the file
def findFile(filename, directoryPath):
    result = None
    path = os.path.join(directoryPath, filename + '.csv')
    if os.path.isfile(path):
        print(path)
        result = path
    else:
        print(path + ' does not exist in directory.')
    return result


# function that standardizes a file based on the column name
# the function locates a specific standardized row (contact_method)
# returns the matched row value of the file
def findContactColumn(place):
    result = None
    dic = matchColumn('Master_List', place)
    for dics in dic:
        # print(dics)
        for key, value in dics.items():
            if key == 'contact_method':
                # print(value)
                result = value
    if result == None:
        print('There is no contact method in this file.')
    else:
        print(result)
    return result


# function that will connect the entire program together
# prints the various results of other functions
# this function acts as the main() of the program
def runProgram():
    place = input('Enter name of place: ')
    path = findDirectory(place)
    if path is not None:
        listFiles(path)
        filename = input('Enter name of file: ')
        filepath = findFile(filename, path)
        if filepath != None:
            column = findContactColumn(place)
            if column != None:
                df = pandas.read_csv(filepath)
                print(df)
                count = df[column].value_counts()
                print(count)


# the program is ran here
print('+---------------------------------------------+')
print('|STANDARDIZATION CONTACT METHOD COUNT PROGRAM |')
print('+---------------------------------------------+')
print('Version 1.0.0')
print('Sukumar Ganpati')
print('Elijah Toussaint')
print('Farzana Yusuf' + '\n')

# path of the master list file that is read and used throughout the program
masterFile = 'Master_List_311_Cities_new (3).xlsx'

# checks to see if the master list file exist
try:
    wb = load_workbook(filename=masterFile)
    ws = wb.active
except FileNotFoundError:
    print('%s not found. Check file or directory...' % masterFile)

# run the program
runProgram()